"""
Utilities for parsing Despacho logs and exporting them to Excel.

The script reads multiple log lines in the format:
<HH:MM> - Abertura da tela de Despacho - <EMP> - EXCEDIDO EM: <XX>%

It detects day transitions based on the operator-provided shift date and
exports the structured data into Excel tables with a maximum of 256
records per block. When the limit is exceeded, new tables are laid out
side by side on the same worksheet. Timestamps are exported with a +3h
offset and a trailing ``Z`` to improve compatibility with
PowerAutomate/SharePoint flows.
"""
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta, time
from pathlib import Path
from typing import Iterable, List

CLOCK_PATTERN = r"\d{2}(?::|h)\d{2}h?"

LOG_PATTERN = re.compile(
    rf"^(?P<clock>{CLOCK_PATTERN})\s*[-–]\s*Abertura da tela de Despacho\s*[-–]\s*"
    rf"(?P<company>[A-Z]{{3}})\s*[-–]\s*EXCEDIDO EM:\s*(?P<percent>\d+)\s*%\s*$"
)


@dataclass
class DispatchRecord:
    """A structured dispatch log entry."""

    timestamp: datetime
    company: str
    exceeded_percent: int

    @property
    def adjusted_iso(self) -> str:
        """Return the timestamp shifted by +3h and suffixed with ``Z``.

        The ISO formatting uses minute precision for compatibility with
        downstream PowerAutomate/SharePoint consumers.
        """

        adjusted = self.timestamp + timedelta(hours=3)
        return adjusted.strftime("%Y-%m-%dT%H:%MZ")


class ParseError(ValueError):
    """Raised when a log line cannot be parsed."""


def _normalize_clock(clock: str) -> str:
    """Normalize times like ``00h00`` or ``00:00h`` to ``HH:MM``."""

    normalized = clock.strip().lower()
    if normalized.endswith("h"):
        normalized = normalized[:-1]
    normalized = normalized.replace("h", ":")

    if not re.fullmatch(r"\d{2}:\d{2}", normalized):
        raise ParseError(f"Horário inválido: {clock!r}")

    return normalized


def parse_records(lines: Iterable[str], shift_date: date) -> List[DispatchRecord]:
    """Parse raw log lines into structured records.

    Args:
        lines: Iterable of log lines in the expected Despacho format.
        shift_date: The operator-provided date marking the start of the shift.

    Returns:
        List of :class:`DispatchRecord` entries with day transitions handled
        whenever the clock time goes backwards (indicating a new day).
    """

    records: List[DispatchRecord] = []
    current_date = shift_date
    previous_clock: time | None = None

    for line_number, raw_line in enumerate(lines, start=1):
        line = raw_line.strip()
        if not line:
            continue

        match = LOG_PATTERN.match(line)
        if not match:
            raise ParseError(
                f"Linha {line_number} não corresponde ao padrão esperado: {raw_line!r}"
            )

        clock_str = match.group("clock")
        company = match.group("company")
        percent = int(match.group("percent"))

        normalized_clock = _normalize_clock(clock_str)
        clock_time = datetime.strptime(normalized_clock, "%H:%M").time()
        if previous_clock and clock_time < previous_clock:
            current_date += timedelta(days=1)

        timestamp = datetime.combine(current_date, clock_time)
        records.append(DispatchRecord(timestamp, company, percent))
        previous_clock = clock_time

    return records


def _record_blocks(records: List[DispatchRecord], block_size: int) -> Iterable[List[DispatchRecord]]:
    for start in range(0, len(records), block_size):
        yield records[start : start + block_size]


def export_to_excel(
    records: List[DispatchRecord],
    output_path: Path,
    max_per_table: int = 256,
) -> None:
    """Export records to an Excel workbook.

    Records are placed in tables of ``max_per_table`` rows. When the number of
    records exceeds this limit, additional tables are placed to the right of the
    previous one on the same worksheet.
    """

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Despacho"

    headers = ["Data/Hora (UTC)", "Empresa", "Excedido (%)"]
    table_width = len(headers)

    for block_index, block in enumerate(_record_blocks(records, max_per_table)):
        start_column = block_index * (table_width + 1) + 1

        # Write headers
        for offset, header in enumerate(headers):
            cell = sheet.cell(row=1, column=start_column + offset)
            cell.value = header

        # Write rows
        for row_offset, record in enumerate(block, start=2):
            sheet.cell(row=row_offset, column=start_column).value = record.adjusted_iso
            sheet.cell(row=row_offset, column=start_column + 1).value = record.company
            sheet.cell(row=row_offset, column=start_column + 2).value = record.exceeded_percent

        # Add a simple border between tables using an empty spacer column header
        spacer_column = get_column_letter(start_column + table_width)
        sheet.column_dimensions[spacer_column].width = 2

    workbook.save(output_path)


def _cli() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Processa linhas de despacho, detecta mudança de dia e exporta para Excel. "
            "Cole as linhas no STDIN ou use --input para apontar um arquivo."
        )
    )
    parser.add_argument(
        "--data",
        dest="shift_date",
        required=True,
        help="Data inicial do turno (formato YYYY-MM-DD)",
        type=date.fromisoformat,
    )
    parser.add_argument(
        "--output",
        dest="output",
        default="despacho_export.xlsx",
        type=Path,
        help="Caminho do arquivo Excel a ser gerado",
    )
    parser.add_argument(
        "--input",
        dest="input_file",
        type=Path,
        help="Arquivo contendo as linhas de despacho. Se omitido, usa STDIN.",
    )
    parser.add_argument(
        "--max-per-table",
        dest="max_per_table",
        type=int,
        default=256,
        help="Limite de linhas por tabela antes de criar uma nova ao lado.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = _cli()
    args = parser.parse_args(argv)

    if args.input_file:
        if not args.input_file.exists():
            parser.error(f"Arquivo de entrada não encontrado: {args.input_file}")
        lines = args.input_file.read_text(encoding="utf-8").splitlines()
    else:
        if sys.stdin.isatty():
            print(
                "Cole as linhas do log e pressione Ctrl+D (Linux/macOS) ou Ctrl+Z (Windows) para finalizar:\n",
                file=sys.stderr,
                end="",
            )
        lines = sys.stdin.read().splitlines()

    records = parse_records(lines, args.shift_date)
    export_to_excel(records, args.output, max_per_table=args.max_per_table)

    print(f"Exportados {len(records)} registros para {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
